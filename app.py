from flask import Flask, jsonify, request, render_template, send_from_directory
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from werkzeug.utils import secure_filename
import os, uuid, re, tempfile, base64, io
import pytesseract
import cv2
import numpy as np
from PIL import Image

# Auto-detect Tesseract: Linux/cloud uses system binary; Windows uses local install
import shutil as _shutil
_tess_system = _shutil.which("tesseract")
if _tess_system:
    pytesseract.pytesseract.tesseract_cmd = _tess_system
else:
    pytesseract.pytesseract.tesseract_cmd = r"C:\Users\mm0225\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app = Flask(__name__,
            template_folder=os.path.join(BASE_DIR, "templates"),
            static_folder=os.path.join(BASE_DIR, "static"))
app.config["MAX_CONTENT_LENGTH"] = 5 * 1024 * 1024   # 5 MB max upload

# Always return JSON errors instead of HTML error pages
@app.errorhandler(Exception)
def handle_exception(e):
    import traceback
    msg = str(e)
    # PermissionError means the Excel file is open in another program
    if isinstance(e, PermissionError):
        msg = "Excel file is open in another program (e.g. Microsoft Excel). Please close it and try again."
    return jsonify({"error": msg, "trace": traceback.format_exc()[-600:]}), 500

@app.errorhandler(404)
def handle_404(e): return jsonify({"error": "Not found"}), 404

@app.errorhandler(413)
def handle_413(e): return jsonify({"error": "File too large (max 5 MB)"}), 413

# DATA_DIR env var lets cloud hosts point to a persistent volume (e.g. /data)
DATA_DIR    = os.environ.get("DATA_DIR", BASE_DIR)
EXCEL_FILE  = os.path.join(DATA_DIR, "Data_Entry_Form.xlsx")
UPLOADS_DIR = os.path.join(DATA_DIR, "uploads")
ALLOWED_EXT = {"png","jpg","jpeg","gif","webp"}

os.makedirs(UPLOADS_DIR, exist_ok=True)

DEFAULT_CLASSES = [
    "1st","2nd","3rd","4th","5th","6th","7th","8th","9th","10th","11th","12th",
    "BA 1st","BA 2nd","BA 3rd","MA 1st Year","MA 2nd Year",
]

def allowed(filename):
    return "." in filename and filename.rsplit(".",1)[1].lower() in ALLOWED_EXT

# ── Excel helpers ─────────────────────────────────────────────────────────────
def hstyle():
    return (Font(bold=True,color="FFFFFF",size=11),
            PatternFill("solid",fgColor="0D1B2A"),
            Alignment(horizontal="center",vertical="center"))

def _apply_hdr(cell, val):
    hf,hfi,ha = hstyle()
    cell.value=val; cell.font=hf; cell.fill=hfi; cell.alignment=ha

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb=openpyxl.Workbook()
        ws1=wb.active; ws1.title="Entry Form"
        _init_data(wb.create_sheet("Data"))
        _init_classes(wb.create_sheet("Classes"))
        wb.save(EXCEL_FILE); return
    wb=openpyxl.load_workbook(EXCEL_FILE); changed=False
    if "Data" not in wb.sheetnames:
        _init_data(wb.create_sheet("Data")); changed=True
    else:
        ws=wb["Data"]; hdr=[c.value for c in ws[1]]
        if "Marks" not in hdr:
            ws.insert_cols(4); _apply_hdr(ws.cell(1,4),"Marks")
            hdr=[c.value for c in ws[1]]; changed=True
        if "Class" not in hdr:
            ws.insert_cols(4); _apply_hdr(ws.cell(1,4),"Class")
            hdr=[c.value for c in ws[1]]; changed=True
        if "Image" not in hdr:
            # add after Marks (col 6) — before Date & Time
            dt_col=next((i+1 for i,v in enumerate(hdr) if v=="Date & Time"),None)
            ins=dt_col if dt_col else len(hdr)+1
            ws.insert_cols(ins); _apply_hdr(ws.cell(1,ins),"Image")
            ws.column_dimensions[chr(64+ins)].width=22; changed=True
    if "Classes" not in wb.sheetnames:
        _init_classes(wb.create_sheet("Classes")); changed=True
    if changed: wb.save(EXCEL_FILE)

def _init_data(ws):
    ws.append(["Sr. No.","Name","Mobile No","Class","Marks","Image","Date & Time"])
    hf,hfi,ha=hstyle()
    for c in ws[1]: c.font=hf; c.fill=hfi; c.alignment=ha
    for col,w in zip("ABCDEFG",[10,25,18,16,12,22,25]):
        ws.column_dimensions[col].width=w

def _init_classes(ws):
    ws.append(["Class Name"])
    _apply_hdr(ws.cell(1,1),"Class Name")
    ws.column_dimensions["A"].width=20
    for c in DEFAULT_CLASSES: ws.append([c])

def _border():
    t=Side(style="thin")
    return Border(left=t,right=t,top=t,bottom=t)

# ── Image route ───────────────────────────────────────────────────────────────
@app.route("/uploads/<path:filename>")
def uploaded_file(filename):
    return send_from_directory(UPLOADS_DIR, filename)

# ── API: Classes ──────────────────────────────────────────────────────────────
@app.route("/api/classes", methods=["GET"])
def get_classes():
    wb=openpyxl.load_workbook(EXCEL_FILE)
    ws=wb["Classes"]
    return jsonify([r[0] for r in ws.iter_rows(min_row=2,values_only=True) if r[0]])

@app.route("/api/classes", methods=["POST"])
def add_class():
    name=(request.json or {}).get("name","").strip()
    if not name: return jsonify({"error":"Name required"}),400
    wb=openpyxl.load_workbook(EXCEL_FILE); ws=wb["Classes"]
    existing=[r[0] for r in ws.iter_rows(min_row=2,values_only=True) if r[0]]
    if name in existing: return jsonify({"error":f'"{name}" already exists'}),409
    r=ws.max_row+1; c=ws.cell(r,1,name)
    c.border=_border(); c.alignment=Alignment(horizontal="center")
    wb.save(EXCEL_FILE)
    return jsonify({"ok":True,"name":name})

@app.route("/api/classes/<path:name>", methods=["DELETE"])
def del_class(name):
    wb=openpyxl.load_workbook(EXCEL_FILE); ws=wb["Classes"]
    target=next((row[0].row for row in ws.iter_rows(min_row=2) if row[0].value==name),None)
    if target is None: return jsonify({"error":"Not found"}),404
    ws.delete_rows(target); wb.save(EXCEL_FILE)
    return jsonify({"ok":True})

# ── API: Records ──────────────────────────────────────────────────────────────
@app.route("/api/records", methods=["GET"])
def get_records():
    wb=openpyxl.load_workbook(EXCEL_FILE); ws=wb["Data"]
    hdr=[c.value for c in ws[1]]
    def idx(col): return hdr.index(col) if col in hdr else None
    rows=[]
    for r in ws.iter_rows(min_row=2,values_only=True):
        if r[0] is None: continue
        def g(col): i=idx(col); return r[i] if i is not None and i<len(r) else None
        rows.append({
            "sr":    r[0],
            "name":  g("Name"),
            "mobile":g("Mobile No"),
            "cls":   g("Class"),
            "marks": g("Marks"),
            "image": g("Image") or "",
            "date":  str(g("Date & Time")) if g("Date & Time") else ""
        })
    return jsonify(rows)

@app.route("/api/records", methods=["POST"])
def add_record():
    # Accept multipart/form-data (fields + optional file)
    name   = request.form.get("name","").strip()
    mobile = request.form.get("mobile","").strip()
    cls    = request.form.get("cls","").strip()
    marks  = request.form.get("marks","").strip()

    if not name:   return jsonify({"error":"Name is required"}),400
    if not mobile: return jsonify({"error":"Mobile is required"}),400
    if not cls:    return jsonify({"error":"Class is required"}),400
    if not marks:  return jsonify({"error":"Marks is required"}),400
    if not mobile.isdigit() or len(mobile)!=10:
        return jsonify({"error":"Mobile must be 10 digits"}),400
    try:
        mv=float(marks)
        if not(0<=mv<=100): raise ValueError
    except: return jsonify({"error":"Marks must be 0–100"}),400

    # Handle image: prefer cropped token, fallback to raw upload
    img_filename=""
    crop_token = request.form.get("crop_token","").strip()
    if crop_token and re.fullmatch(r"[0-9a-f]{16}", crop_token):
        tmp_path = os.path.join(tempfile.gettempdir(), f"crop_{crop_token}.png")
        if os.path.exists(tmp_path):
            img_filename = f"student_{uuid.uuid4().hex[:12]}.png"
            dest = os.path.join(UPLOADS_DIR, img_filename)
            import shutil; shutil.move(tmp_path, dest)
    if not img_filename:
        f=request.files.get("image")
        if f and f.filename and allowed(f.filename):
            ext=f.filename.rsplit(".",1)[1].lower()
            img_filename=f"student_{uuid.uuid4().hex[:12]}.{ext}"
            f.save(os.path.join(UPLOADS_DIR,img_filename))

    try:
        wb=openpyxl.load_workbook(EXCEL_FILE); ws=wb["Data"]
        row=ws.max_row+1; serial=row-1
        ts=datetime.now().strftime("%d-%b-%Y %H:%M:%S")
        border=_border()
        for col,val in enumerate([serial,name,mobile,cls,mv,img_filename,ts],1):
            c=ws.cell(row,col,val); c.border=border
            c.alignment=Alignment(horizontal="left" if col==2 else "center")
        wb.save(EXCEL_FILE)
    except PermissionError:
        return jsonify({"error": "Cannot save — Data_Entry_Form.xlsx is open in Excel. Close it and try again."}), 500
    except Exception as ex:
        return jsonify({"error": f"Save failed: {ex}"}), 500
    return jsonify({"ok":True,"serial":serial,"image":img_filename})

@app.route("/api/records/<int:serial_no>", methods=["DELETE"])
def del_record(serial_no):
    wb=openpyxl.load_workbook(EXCEL_FILE); ws=wb["Data"]
    hdr=[c.value for c in ws[1]]
    img_col=hdr.index("Image") if "Image" in hdr else None
    target=None; img_file=None
    for row in ws.iter_rows(min_row=2):
        if row[0].value is not None and int(row[0].value)==serial_no:
            target=row[0].row
            if img_col is not None: img_file=row[img_col].value
            break
    if target is None: return jsonify({"error":"Not found"}),404
    ws.delete_rows(target)
    border=_border()
    for i,row in enumerate(ws.iter_rows(min_row=2),1):
        if row[0].value is not None:
            row[0].value=i
            for cell in row: cell.border=border
    wb.save(EXCEL_FILE)
    # Delete image file if it exists
    if img_file:
        fp=os.path.join(UPLOADS_DIR,img_file)
        if os.path.exists(fp): os.remove(fp)
    return jsonify({"ok":True})

@app.route("/api/crop", methods=["POST"])
def crop_face():
    """
    Receive an ID-card image, detect the face region, crop it with padding,
    and return the result as a base64 PNG so the frontend can preview it.
    Also saves a temp file and returns its token so the save-record call can
    retrieve the already-cropped image without re-uploading the full card.
    """
    f = request.files.get("image")
    if not f or not f.filename:
        return jsonify({"error": "No image provided"}), 400

    # Decode with OpenCV
    file_bytes = np.frombuffer(f.read(), np.uint8)
    img_bgr = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)
    if img_bgr is None:
        return jsonify({"error": "Cannot decode image"}), 400

    # Enhance contrast for low-quality ID scans
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    gray = cv2.equalizeHist(gray)

    # Try cascades from strictest → most permissive; stop at first hit
    cascade_dir = cv2.data.haarcascades
    faces = []
    attempts = [
        ("haarcascade_frontalface_default.xml", 1.1,  4, (40,40)),
        ("haarcascade_frontalface_default.xml", 1.05, 3, (30,30)),
        ("haarcascade_frontalface_alt.xml",     1.05, 3, (30,30)),
        ("haarcascade_frontalface_alt2.xml",    1.05, 3, (30,30)),
        ("haarcascade_frontalface_default.xml", 1.03, 2, (25,25)),
        ("haarcascade_profileface.xml",         1.05, 3, (30,30)),
    ]
    for cascade_file, scale, neighbors, min_sz in attempts:
        path = os.path.join(cascade_dir, cascade_file)
        if not os.path.exists(path):
            continue
        cascade = cv2.CascadeClassifier(path)
        faces = cascade.detectMultiScale(gray, scaleFactor=scale,
                                         minNeighbors=neighbors, minSize=min_sz)
        if len(faces):
            break

    if not len(faces):
        # Return the original image as preview so UI still shows something
        ok, buf = cv2.imencode(".png", img_bgr)
        b64 = base64.b64encode(buf.tobytes()).decode()
        return jsonify({
            "found": False,
            "preview": f"data:image/png;base64,{b64}",
            "message": "No face detected — full image will be stored. For best results upload a clear front-facing photo."
        })

    # Pick the largest face (most prominent on ID card)
    x, y, w, h = max(faces, key=lambda r: r[2] * r[3])

    # Add padding (40% each side) so we don't clip hair/chin
    ih, iw = img_bgr.shape[:2]
    pad_x = int(w * 0.45)
    pad_y = int(h * 0.55)
    x1 = max(0, x - pad_x)
    y1 = max(0, y - pad_y)
    x2 = min(iw, x + w + pad_x)
    y2 = min(ih, y + h + pad_y)

    cropped = img_bgr[y1:y2, x1:x2]

    # Encode cropped image to PNG → base64 for preview
    ok, buf = cv2.imencode(".png", cropped)
    b64 = base64.b64encode(buf.tobytes()).decode()

    # Save cropped image to a temp file so save-record can use it
    token = uuid.uuid4().hex[:16]
    tmp_path = os.path.join(tempfile.gettempdir(), f"crop_{token}.png")
    cv2.imwrite(tmp_path, cropped)

    return jsonify({
        "found": True,
        "token": token,
        "preview": f"data:image/png;base64,{b64}"
    })


@app.route("/api/crop/<token>", methods=["GET"])
def get_cropped(token):
    """Return the previously cropped temp file as a PNG download."""
    if not re.fullmatch(r"[0-9a-f]{16}", token):
        return jsonify({"error": "Invalid token"}), 400
    tmp_path = os.path.join(tempfile.gettempdir(), f"crop_{token}.png")
    if not os.path.exists(tmp_path):
        return jsonify({"error": "Expired or not found"}), 404
    return send_from_directory(tempfile.gettempdir(), f"crop_{token}.png",
                               mimetype="image/png")


@app.route("/api/ocr", methods=["POST"])
def ocr_image():
    f = request.files.get("image")
    if not f or not f.filename:
        return jsonify({"error": "No image provided"}), 400

    # Save to temp file for Tesseract
    ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else "png"
    tmp_path = os.path.join(tempfile.gettempdir(), f"ocr_tmp_{uuid.uuid4().hex[:8]}.{ext}")
    try:
        f.save(tmp_path)
        img = Image.open(tmp_path)
        # Upscale small images for better accuracy
        w, h = img.size
        if w < 800:
            scale = 800 / w
            img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
        raw_text = pytesseract.image_to_string(img, lang="eng", config="--psm 6")
    finally:
        try: os.remove(tmp_path)
        except: pass

    extracted = _parse_ocr(raw_text)
    return jsonify({"raw": raw_text.strip(), "fields": extracted})


def _parse_ocr(text):
    """
    Try to pull Name / Mobile / Class / Marks from raw OCR text.
    Looks for labelled lines first, then falls back to pattern matching.
    """
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    result = {"name": "", "mobile": "", "cls": "", "marks": ""}

    # ── Mobile: 10-digit number (Indian pattern first) ────────────────
    mobile_m = re.search(r"\b([6-9]\d{9})\b", text)
    if not mobile_m:
        mobile_m = re.search(r"\b(\d{10})\b", text)
    if mobile_m:
        result["mobile"] = mobile_m.group(1)

    # ── Marks: number after keyword, else first standalone 0–100 int ──
    marks_m = re.search(r"(?:marks?|score|total|percentage)\s*[:\-=]\s*(\d{1,3}(?:\.\d+)?)", text, re.I)
    if not marks_m:
        marks_m = re.search(r"(?:marks?|score|total|percentage)\s+(\d{1,3}(?:\.\d+)?)", text, re.I)
    if not marks_m:
        for m in re.finditer(r"\b(\d{1,3})\b", text):
            v = int(m.group(1))
            if 0 <= v <= 100 and m.group(1) != result["mobile"]:
                marks_m = m; break
    if marks_m:
        result["marks"] = marks_m.group(1)

    # ── Class: ordered longest-match first to prevent "1st" eating "10th" ──
    class_keywords = [
        # Higher classes checked BEFORE lower ones to avoid prefix collisions
        ("12th",       [r"\b12th\b", r"\bclass\s*12\b", r"\bgrade\s*12\b", r"\bxii\b"]),
        ("11th",       [r"\b11th\b", r"\bclass\s*11\b", r"\bgrade\s*11\b", r"\bxi\b"]),
        ("10th",       [r"\b10th\b", r"\bclass\s*10\b", r"\bgrade\s*10\b", r"\bx\b"]),
        ("MA 2nd Year",[r"\bma\s*2\b", r"\bma\s*ii\b", r"\bma\s*second\b", r"\bmaster.*2\b"]),
        ("MA 1st Year",[r"\bma\s*1\b", r"\bma\s*i\b",  r"\bma\s*first\b",  r"\bmaster.*1\b"]),
        ("BA 3rd",     [r"\bba\s*3\b", r"\bba\s*iii\b", r"\bba\s*third\b"]),
        ("BA 2nd",     [r"\bba\s*2\b", r"\bba\s*ii\b",  r"\bba\s*second\b"]),
        ("BA 1st",     [r"\bba\s*1\b", r"\bba\s*i\b",   r"\bba\s*first\b", r"\bbachelor.*1\b"]),
        ("9th",        [r"\b9th\b",  r"\bclass\s*9\b",  r"\bgrade\s*9\b",  r"\bix\b"]),
        ("8th",        [r"\b8th\b",  r"\bclass\s*8\b",  r"\bgrade\s*8\b",  r"\bviii\b"]),
        ("7th",        [r"\b7th\b",  r"\bclass\s*7\b",  r"\bgrade\s*7\b",  r"\bvii\b"]),
        ("6th",        [r"\b6th\b",  r"\bclass\s*6\b",  r"\bgrade\s*6\b",  r"\bvi\b"]),
        ("5th",        [r"\b5th\b",  r"\bclass\s*5\b",  r"\bgrade\s*5\b",  r"\bv\b"]),
        ("4th",        [r"\b4th\b",  r"\bclass\s*4\b",  r"\bgrade\s*4\b",  r"\biv\b"]),
        ("3rd",        [r"\b3rd\b",  r"\bclass\s*3\b",  r"\bgrade\s*3\b",  r"\biii\b"]),
        ("2nd",        [r"\b2nd\b",  r"\bclass\s*2\b",  r"\bgrade\s*2\b",  r"\bii\b"]),
        ("1st",        [r"\b1st\b",  r"\bclass\s*1\b",  r"\bgrade\s*1\b",  r"\bstd\s*i\b"]),
    ]
    tl = text.lower()
    for cls_name, patterns in class_keywords:
        for pat in patterns:
            if re.search(pat, tl):
                result["cls"] = cls_name; break
        if result["cls"]: break

    # ── Name: labelled "Name:" / "Student Name:" line ──────────────────
    # Matches: "Student Name: Rahul Sharma"  or  "Name: Rahul Sharma"
    name_m = re.search(
        r"(?:student\s+name|full\s+name|name)\s*[:\-=]\s*([A-Za-z][A-Za-z\s\.]{2,50})",
        text, re.I
    )
    if name_m:
        # strip any trailing label-like words ("mobile","class","marks")
        val = re.split(r"\b(?:mobile|class|grade|marks?|score|roll|dob|date)\b", name_m.group(1), flags=re.I)[0]
        result["name"] = val.strip().title()
    else:
        # Fallback: first line with 2+ alphabetic words, no digits
        for line in lines:
            clean = re.sub(r"[^A-Za-z\s]", "", line).strip()
            words = [w for w in clean.split() if len(w) > 1]
            if len(words) >= 2 and not re.search(r"\d", line):
                result["name"] = clean.title(); break

    return result


@app.route("/")
def index(): return render_template("index.html")

if __name__=="__main__":
    init_excel()
    print("\n  ✅  FIFA 2026 Data Entry — open http://127.0.0.1:5000\n")
    app.run(debug=False,port=5000)
